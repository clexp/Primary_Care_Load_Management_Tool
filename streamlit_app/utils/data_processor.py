import pandas as pd
import glob
import os
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import logging
from datetime import datetime

# Set up logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# Define the directory path relative to the script location
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(SCRIPT_DIR, "Attachments-call data")

# Required columns for call data
REQUIRED_COLUMNS = ['Time', 'Total Calls']

def create_call_data_template():
    """Create a template Excel file for call data"""
    # Create a new workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Call Data"
    
    # Define headers
    headers = ["Time", "Total Calls", "Answered Calls", "Abandoned Calls", "Avg Wait Time (sec)", "Avg Talk Time (sec)"]
    
    # Add headers with styling
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    
    # Add sample data (one day with hourly data)
    sample_data = [
        ["2023-01-01 08:00:00", 5, 4, 1, 45, 180],
        ["2023-01-01 09:00:00", 8, 7, 1, 60, 210],
        ["2023-01-01 10:00:00", 12, 10, 2, 90, 240],
        ["2023-01-01 11:00:00", 15, 13, 2, 120, 270],
        ["2023-01-01 12:00:00", 10, 9, 1, 75, 225],
        ["2023-01-01 13:00:00", 7, 6, 1, 50, 195],
        ["2023-01-01 14:00:00", 9, 8, 1, 65, 210],
        ["2023-01-01 15:00:00", 11, 10, 1, 80, 240],
        ["2023-01-01 16:00:00", 6, 5, 1, 40, 180],
        ["2023-01-01 17:00:00", 4, 4, 0, 30, 150],
    ]
    
    for row_idx, row_data in enumerate(sample_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(horizontal="center")
    
    # Add instructions sheet
    ws_instructions = wb.create_sheet("Instructions")
    ws_instructions['A1'] = "Call Data Template Instructions"
    ws_instructions['A1'].font = Font(bold=True, size=14)
    
    instructions = [
        "This template shows the required format for call data files:",
        "",
        "1. Time: Date and time in format 'YYYY-MM-DD HH:MM:SS'",
        "2. Total Calls: Total number of calls received in the time period",
        "3. Answered Calls: Number of calls that were answered",
        "4. Abandoned Calls: Number of calls that were abandoned",
        "5. Avg Wait Time (sec): Average wait time in seconds",
        "6. Avg Talk Time (sec): Average talk time in seconds",
        "",
        "Notes:",
        "- You can add more rows following the same format",
        "- The Time column must be in the correct datetime format",
        "- All numeric columns should contain numbers only",
        "- Save your file as CSV before uploading to the application"
    ]
    
    for i, instruction in enumerate(instructions, 3):
        ws_instructions[f'A{i}'] = instruction
    
    # Adjust column widths
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 20
    
    # Save to bytes
    from io import BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output

def create_staffing_template():
    """Create a template Excel file for staffing data"""
    # Create a new workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Staffing Levels"
    
    # Define headers
    time_slots = [f"{hour:02d}:00" for hour in range(8, 18)] + [f"{hour:02d}:30" for hour in range(8, 18)]
    time_slots.sort()
    days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]
    
    # Add headers with styling
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # Add time column header
    cell = ws.cell(row=1, column=1, value="Time")
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center")
    
    # Add day headers
    for col, day in enumerate(days, 2):
        cell = ws.cell(row=1, column=col, value=day)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    
    # Add time slots
    for row, time_slot in enumerate(time_slots, 2):
        ws.cell(row=row, column=1, value=time_slot)
        ws.cell(row=row, column=1).alignment = Alignment(horizontal="center")
    
    # Add sample data (some cells with values)
    sample_data = {
        "09:00": {"Monday": 5, "Wednesday": 6},
        "10:30": {"Tuesday": 7, "Thursday": 8},
        "14:00": {"Monday": 6, "Friday": 5},
        "16:30": {"Tuesday": 4, "Wednesday": 5}
    }
    
    for time_slot, day_data in sample_data.items():
        row = time_slots.index(time_slot) + 2
        for day, value in day_data.items():
            col = days.index(day) + 2
            cell = ws.cell(row=row, column=col, value=value)
            cell.alignment = Alignment(horizontal="center")
    
    # Add instructions sheet
    ws_instructions = wb.create_sheet("Instructions")
    ws_instructions['A1'] = "Staffing Levels Template Instructions"
    ws_instructions['A1'].font = Font(bold=True, size=14)
    
    instructions = [
        "This template shows the required format for staffing level data:",
        "",
        "1. Time: Time slots in 30-minute intervals from 8:00 to 17:30",
        "2. Days: Monday through Friday",
        "3. Values: Number of staff members scheduled for each time slot and day",
        "",
        "Notes:",
        "- Enter the number of staff members in each cell",
        "- Leave cells blank or enter 0 for time slots with no staff",
        "- The application will use this data to calculate staffing gaps",
        "- Save your file as CSV before uploading to the application"
    ]
    
    for i, instruction in enumerate(instructions, 3):
        ws_instructions[f'A{i}'] = instruction
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 15
    for col in range(2, len(days) + 2):
        ws.column_dimensions[get_column_letter(col)].width = 12
    
    # Save to bytes
    from io import BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output

def validate_dataframe(df, file_name):
    """
    Validate a dataframe to ensure it has the required columns and data types
    
    Args:
        df: DataFrame to validate
        file_name: Name of the file being validated (for error messages)
        
    Returns:
        tuple: (is_valid, error_message)
    """
    # Check for required columns
    missing_columns = [col for col in REQUIRED_COLUMNS if col not in df.columns]
    if missing_columns:
        return False, f"Missing required columns: {', '.join(missing_columns)}"
    
    # Check for empty dataframe
    if df.empty:
        return False, "File contains no data"
    
    # Check for non-numeric values in Total Calls column and log a warning
    non_numeric_mask = ~pd.to_numeric(df['Total Calls'], errors='coerce').notna()
    if non_numeric_mask.any():
        non_numeric_count = non_numeric_mask.sum()
        non_numeric_examples = df.loc[non_numeric_mask, 'Total Calls'].head(3).tolist()
        logger.warning(f"Found {non_numeric_count} non-numeric values in Total Calls column. Examples: {non_numeric_examples}")
        
        # Clean up non-numeric values by replacing them with NaN
        df.loc[non_numeric_mask, 'Total Calls'] = np.nan
        
        # Log the number of rows affected
        logger.info(f"Cleaned {non_numeric_count} non-numeric values in Total Calls column")
    
    return True, ""

def detect_outliers(df, column, threshold=3):
    """
    Detect outliers in a column using the z-score method
    
    Args:
        df: DataFrame containing the data
        column: Column name to check for outliers
        threshold: Z-score threshold for outlier detection
        
    Returns:
        DataFrame: DataFrame with outliers marked
    """
    # Calculate z-score
    z_scores = np.abs((df[column] - df[column].mean()) / df[column].std())
    
    # Mark outliers
    df[f'{column}_is_outlier'] = z_scores > threshold
    
    return df

def process_call_data(file):
    """
    Process uploaded call data file with validation and error handling
    
    Args:
        file: Uploaded file object
        
    Returns:
        tuple: (DataFrame, dict) - Processed DataFrame and data quality metrics
    """
    try:
        # Read the file based on its extension
        if file.name.endswith('.csv'):
            df = pd.read_csv(file)
        elif file.name.endswith('.xlsx'):
            df = pd.read_excel(file)
        else:
            raise ValueError(f"Unsupported file format: {file.name}")
        
        # Validate the dataframe
        is_valid, error_message = validate_dataframe(df, file.name)
        if not is_valid:
            raise ValueError(error_message)
        
        # Remove the "Total" row if present
        if 'Time' in df.columns and df['Time'].astype(str).str.contains('Total').any():
            df = df[~df['Time'].astype(str).str.contains('Total')].copy()
        
        # Convert Time column to datetime
        try:
            df['Time'] = pd.to_datetime(df['Time'])
        except Exception as e:
            raise ValueError(f"Error converting Time column to datetime: {str(e)}")
        
        # Handle NaN values in Total Calls column
        if df['Total Calls'].isna().any():
            # Log the number of NaN values
            nan_count = df['Total Calls'].isna().sum()
            logger.info(f"Found {nan_count} NaN values in Total Calls column")
            
            # Fill NaN values with 0 or with the mean of the column
            # Using 0 as a default, but you could use df['Total Calls'].mean() or other methods
            df['Total Calls'] = df['Total Calls'].fillna(0)
            logger.info(f"Filled {nan_count} NaN values in Total Calls column with 0")
        
        # Calculate data quality metrics
        quality_metrics = calculate_data_quality_metrics(df)
        
        # Detect outliers in Total Calls column
        df = detect_outliers(df, 'Total Calls')
        
        return df, quality_metrics
    
    except Exception as e:
        logger.error(f"Error processing file {file.name}: {str(e)}")
        raise

def calculate_data_quality_metrics(df):
    """
    Calculate data quality metrics for a DataFrame
    
    Args:
        df: DataFrame to analyze
        
    Returns:
        dict: Dictionary of data quality metrics
    """
    metrics = {
        'total_records': len(df),
        'missing_values': df.isnull().sum().to_dict(),
        'date_range': {
            'start': df['Time'].min().strftime('%Y-%m-%d %H:%M:%S') if not df.empty else None,
            'end': df['Time'].max().strftime('%Y-%m-%d %H:%M:%S') if not df.empty else None
        },
        'total_calls': df['Total Calls'].sum() if 'Total Calls' in df.columns else 0,
        'avg_calls_per_record': df['Total Calls'].mean() if 'Total Calls' in df.columns else 0,
        'max_calls': df['Total Calls'].max() if 'Total Calls' in df.columns else 0,
        'min_calls': df['Total Calls'].min() if 'Total Calls' in df.columns else 0,
        'outliers': {
            'count': df['Total_Calls_is_outlier'].sum() if 'Total_Calls_is_outlier' in df.columns else 0
        }
    }
    
    return metrics

def combine_call_data(dataframes):
    """
    Combine multiple call data DataFrames with validation
    
    Args:
        dataframes: List of DataFrames to combine
        
    Returns:
        tuple: (DataFrame, dict) - Combined DataFrame and data quality metrics
    """
    if not dataframes:
        raise ValueError("No dataframes provided")
    
    # Combine all dataframes
    combined_df = pd.concat(dataframes, ignore_index=True)
    
    # Sort by time to ensure chronological order
    combined_df = combined_df.sort_values('Time')
    
    # Remove duplicates if any
    combined_df = combined_df.drop_duplicates(subset=['Time'])
    
    # Calculate data quality metrics for the combined dataframe
    quality_metrics = calculate_data_quality_metrics(combined_df)
    
    return combined_df, quality_metrics

def load_and_combine_call_data(directory=DATA_DIR):
    """
    Load and combine call data from CSV files in a directory
    
    Args:
        directory: Directory containing CSV files
        
    Returns:
        DataFrame: Combined DataFrame
    """
    # Check if directory exists
    if not os.path.exists(directory):
        raise FileNotFoundError(f"Directory not found: {directory}")
    
    # Get all CSV files in the directory that match the pattern
    csv_files = glob.glob(os.path.join(directory, "csv-queue-report*.csv"))
    
    if not csv_files:
        raise FileNotFoundError(f"No CSV files found in {directory}")
    
    # List to store individual dataframes
    dfs = []
    
    # Read each CSV file
    for file in csv_files:
        logger.info(f"Reading file: {os.path.basename(file)}")
        try:
            # Read CSV file
            df = pd.read_csv(file)
            
            # Validate the dataframe
            is_valid, error_message = validate_dataframe(df, os.path.basename(file))
            if not is_valid:
                logger.warning(f"Skipping file {os.path.basename(file)}: {error_message}")
                continue
            
            # Remove the "Total" row
            df = df[df['Time'] != 'Total']
            
            # Convert Time column to datetime
            df['Time'] = pd.to_datetime(df['Time'])
            
            # Add to list of dataframes
            dfs.append(df)
        except Exception as e:
            logger.error(f"Error processing file {os.path.basename(file)}: {str(e)}")
    
    if not dfs:
        raise ValueError("No valid dataframes found")
    
    # Combine all dataframes
    combined_df, quality_metrics = combine_call_data(dfs)
    
    return combined_df

def load_demo_data():
    """Load synthetic demo data"""
    # Create synthetic data similar to your real data
    # This can be based on your current dataset
    return demo_df

if __name__ == "__main__":
    try:
        # Load and combine the data
        combined_data = load_and_combine_call_data()
        
        # Print basic information about the combined dataset
        print(f"\nTotal number of records: {len(combined_data)}")
        print(f"\nDate range: from {combined_data['Time'].min()} to {combined_data['Time'].max()}")
        print(f"\nColumns in the dataset:")
        for col in combined_data.columns:
            print(f"- {col}")
    except Exception as e:
        print(f"Error: {e}") 